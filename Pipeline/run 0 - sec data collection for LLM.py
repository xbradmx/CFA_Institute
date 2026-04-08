"""
DDDS - SEC Filing Collection Pipeline
======================================
Downloads the most recent original 10-K and 10-Q filings from SEC EDGAR
for all US-listed companies in SIC 3400–3599 (Fabricated Metals &
Industrial/Commercial Machinery). This is the entry point of the DDDS
pipeline — all downstream scoring, graph-building, and memo generation
depend on the filings collected here.

What this script does
---------------------
1. Builds the company universe by scanning every SEC-registered company
   via the EDGAR submissions API, filtering to SIC 3400–3599. Results
   are cached after the first run (~19 minutes) so subsequent runs are
   instant.
2. Excludes pure-OTC listings, foreign private issuers (20-F / 6-K
   filers), and duplicate share classes (one ticker per CIK).
3. For each company, downloads the 2 most recent 10-K filings and the
   3 most recent 10-Q filings within the collection window.
4. Amendments (10-K/A, 10-Q/A) are completely ignored — they lack the
   structured section headings required for downstream NLP extraction.
5. Generates a summary Excel workbook showing per-company coverage.

Collection window
-----------------
Start:  2024-01-01
End:    2025-09-30
This 21-month window is wide enough to capture two annual filing cycles
and at least three quarterly filings for every company.

Directory structure created
---------------------------
    data/Filings by company/
        {TICKER}/
            10-K/
                {TICKER}_10-K_{DATE}_{ACCESSION}.html
            10-Q/
                {TICKER}_10-Q_{DATE}_{ACCESSION}.html

Output
------
    company_filing_summary.xlsx

Usage
-----
    python "run_0_-_sec_data_collection.py"
    python "run_0_-_sec_data_collection.py" --use-cached

Notes
-----
- SEC EDGAR requires a valid User-Agent header with a contact email.
  Update USER_AGENT below before first run.
- The --use-cached flag skips re-downloading filings that already exist
  on disk AND reuses the cached universe JSON. This is required for
  CFA competition reproducibility (API cost < $20 threshold).
- SEC rate limit is 10 requests/second. We target ~6/s to be safe.
"""

import argparse
import json
import os
import re
import time
import logging
from pathlib import Path
from collections import defaultdict

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
USER_AGENT = "DDDS-Research lancaster-university connor@lancaster.ac.uk"

WINDOW_START = "2023-07-01"
WINDOW_END   = "2025-09-30"

TARGET_10K = 2
TARGET_10Q = 3

# SIC range: Fabricated Metal Products (3400–3499) +
#             Industrial & Commercial Machinery (3500–3599)
SIC_START = 3400
SIC_END   = 3599

# Relative to wherever this script is run from (Pipeline/)
FILINGS_DIR    = Path("data/Filings by company")
UNIVERSE_CACHE = Path("data/company_universe_cache.json")
EXCEL_OUTPUT   = Path("company_filing_summary.xlsx")
LOG_FILE       = Path("data/sec_collection.log")

EDGAR_HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept-Encoding": "gzip, deflate",
}
REQUEST_DELAY = 0.15  # ~6 req/s, well under SEC's 10/s limit
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------
def setup_logging():
    os.makedirs("data", exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(str(LOG_FILE)),
            logging.StreamHandler(),
        ],
    )
    return logging.getLogger("ddds-collector")


# ---------------------------------------------------------------------------
# HTTP HELPERS
# ---------------------------------------------------------------------------
_delay = REQUEST_DELAY
_request_count = 0


def rate_limit():
    global _delay
    time.sleep(_delay)
    # Decay back toward base delay after successful requests
    if _delay > REQUEST_DELAY:
        _delay = max(REQUEST_DELAY, _delay * 0.995)


def safe_get(url: str, max_retries: int = 3):
    """GET with retries, adaptive backoff on 429s."""
    global _delay, _request_count
    for attempt in range(max_retries):
        try:
            rate_limit()
            _request_count += 1
            r = requests.get(url, headers=EDGAR_HEADERS, timeout=30)
            if r.status_code == 200:
                return r
            if r.status_code == 429:
                _delay = min(_delay * 2, 5.0)
                wait = 15 * (attempt + 1)
                logging.warning(
                    f"Rate limited (429). Delay → {_delay:.2f}s. "
                    f"Sleeping {wait}s..."
                )
                time.sleep(wait)
                continue
            if r.status_code == 404:
                return None
            logging.warning(f"HTTP {r.status_code} for {url}")
            time.sleep(3)
        except requests.exceptions.RequestException as e:
            logging.warning(f"Request error (attempt {attempt+1}): {e}")
            time.sleep(5)
    return None


# ---------------------------------------------------------------------------
# STEP 1 — BUILD COMPANY UNIVERSE
# ---------------------------------------------------------------------------
def build_company_universe(log) -> dict:
    """
    Scan every SEC-registered company's SIC code via the submissions API.
    This is the only reliable method — EDGAR's browse-by-SIC search index
    is incomplete and misses ~60% of companies.

    Takes ~19 minutes for ~13k companies at 0.12s per request.
    Results are cached to UNIVERSE_CACHE after first run.

    Filtering applied:
        1. SIC 3400–3599 only
        2. Exclude pure-OTC (all exchanges == "OTC")
        3. Exclude foreign private issuers (category contains "foreign")
        4. Deduplicate by CIK (first ticker per CIK wins — skips
           warrants, preferred shares, dual classes like GEF-B, POWWP)

    Returns: dict keyed by CIK string.
    """
    if UNIVERSE_CACHE.exists():
        log.info(f"Loading cached universe from {UNIVERSE_CACHE}")
        with open(UNIVERSE_CACHE) as f:
            cached = json.load(f)
        log.info(f"Loaded {len(cached)} companies from cache.")
        return cached

    os.makedirs(UNIVERSE_CACHE.parent, exist_ok=True)

    # Download full company tickers list
    log.info("Downloading SEC company tickers list...")
    resp = safe_get("https://www.sec.gov/files/company_tickers.json")
    if not resp:
        log.error("FATAL: Cannot download company tickers. Check network.")
        return {}

    all_tickers = resp.json()
    total = len(all_tickers)
    log.info(f"Total SEC-registered entities to scan: {total}")
    log.info(f"Estimated time: ~{total * 0.15 / 60:.0f} minutes")

    companies = {}  # Keyed by CIK — naturally deduplicates share classes
    skipped_otc = 0
    skipped_foreign = 0
    skipped_dupes = 0
    errors = 0

    for i, (key, company) in enumerate(all_tickers.items()):
        cik = str(company["cik_str"])
        cik_padded = cik.zfill(10)
        ticker = company.get("ticker", "")

        # CIK dedup: skip if we already have this company
        if cik in companies:
            skipped_dupes += 1
            continue

        try:
            url = f"https://data.sec.gov/submissions/CIK{cik_padded}.json"
            rate_limit()
            r = requests.get(url, headers=EDGAR_HEADERS, timeout=10)
            if r.status_code != 200:
                errors += 1
                continue

            data = r.json()
            sic = int(data.get("sic", 0) or 0)

            if SIC_START <= sic <= SIC_END:
                # --- Exclusion 1: Pure OTC ---
                exchanges = [e for e in data.get("exchanges", []) if e]
                if exchanges and all(e.upper() == "OTC" for e in exchanges):
                    skipped_otc += 1
                    continue

                # --- Exclusion 2: Foreign private issuers ---
                category = (data.get("category", "") or "").lower()
                if "foreign" in category:
                    skipped_foreign += 1
                    continue

                # Sub-sector label
                sector = ("Fabricated Metal Products" if sic < 3500
                          else "Industrial & Commercial Machinery")

                exchange_str = " | ".join(exchanges) if exchanges else "Unknown"

                companies[cik] = {
                    "cik": cik,
                    "ticker": ticker,
                    "name": data.get("name", ""),
                    "sic": str(sic),
                    "sic_description": data.get("sicDescription", ""),
                    "sector": sector,
                    "exchange": exchange_str,
                    "category": data.get("category", ""),
                }
                log.info(
                    f"  FOUND: {ticker:8s} | SIC {sic} | "
                    f"{data.get('sicDescription', '')}"
                )

        except Exception:
            errors += 1

        # Progress update every 1000 companies
        if (i + 1) % 1000 == 0:
            log.info(
                f"  ...scanned {i+1}/{total} | found {len(companies)} | "
                f"{skipped_dupes} dupes | {skipped_otc} OTC | "
                f"{skipped_foreign} foreign | {errors} errors"
            )

    log.info(
        f"\nUniverse scan complete: {len(companies)} companies "
        f"(SIC {SIC_START}–{SIC_END})"
    )
    log.info(
        f"  Duplicates skipped: {skipped_dupes}  |  "
        f"Pure-OTC excluded: {skipped_otc}  |  "
        f"Foreign issuers excluded: {skipped_foreign}  |  "
        f"Errors: {errors}"
    )

    # Cache results
    with open(UNIVERSE_CACHE, "w") as f:
        json.dump(companies, f, indent=2)
    log.info(f"Cached universe to {UNIVERSE_CACHE}")

    return companies


# ---------------------------------------------------------------------------
# STEP 2 — QUERY EDGAR FOR ORIGINAL FILINGS
# ---------------------------------------------------------------------------
def get_original_filings(cik: str, form_type: str, max_count: int) -> list[dict]:
    """
    Query EDGAR submissions API for the most recent ORIGINAL filings.

    Key behaviour:
        - Exact form match only: "10-K" matches "10-K", NOT "10-K/A"
        - Sorted by filing date descending (most recent first)
        - Returns at most max_count filings within the collection window

    Returns list of filing dicts with: form, date, accession, primary_doc, url
    """
    cik_padded = str(cik).zfill(10)
    url = f"https://data.sec.gov/submissions/CIK{cik_padded}.json"
    resp = safe_get(url)
    if not resp:
        return []

    try:
        data = resp.json()
    except (json.JSONDecodeError, ValueError):
        return []

    recent = data.get("filings", {}).get("recent", {})
    if not recent:
        return []

    forms        = recent.get("form", [])
    dates        = recent.get("filingDate", [])
    accessions   = recent.get("accessionNumber", [])
    primary_docs = recent.get("primaryDocument", [])

    n = min(len(forms), len(dates), len(accessions), len(primary_docs))

    matches = []
    for i in range(n):
        # EXACT match — "10-K" only, never "10-K/A"
        if forms[i] != form_type:
            continue
        if dates[i] < WINDOW_START or dates[i] > WINDOW_END:
            continue

        acc_fmt  = accessions[i]
        acc_path = acc_fmt.replace("-", "")

        filing_url = (
            f"https://www.sec.gov/Archives/edgar/data/"
            f"{cik_padded}/{acc_path}/{primary_docs[i]}"
        )
        matches.append({
            "form":        forms[i],
            "date":        dates[i],
            "accession":   acc_fmt,
            "primary_doc": primary_docs[i],
            "url":         filing_url,
        })

    # Most recent first, take top N
    matches.sort(key=lambda x: x["date"], reverse=True)
    return matches[:max_count]


# ---------------------------------------------------------------------------
# STEP 3 — DOWNLOAD FILINGS
# ---------------------------------------------------------------------------
def filing_extension(primary_doc: str) -> str:
    p = primary_doc.lower()
    if p.endswith(".htm") or p.endswith(".html"):
        return ".html"
    if p.endswith(".txt"):
        return ".txt"
    if p.endswith(".xml"):
        return ".xml"
    return ".html"


def sanitize_ticker(ticker: str) -> str:
    """Clean ticker for use as a folder name."""
    return re.sub(r"[^A-Za-z0-9\-.]", "", ticker).upper()


def download_filing(
    filing: dict, ticker: str, save_dir: Path, use_cached: bool
) -> tuple[str, bool]:
    """
    Download a single filing to disk.
    Skips if file already exists and use_cached is True.
    Returns (filename, success).
    """
    os.makedirs(save_dir, exist_ok=True)

    ext   = filing_extension(filing["primary_doc"])
    fname = f"{ticker}_{filing['form']}_{filing['date']}_{filing['accession']}{ext}"
    fpath = save_dir / fname

    if use_cached and fpath.exists():
        return fname, True

    resp = safe_get(filing["url"])
    if resp and len(resp.content) > 100:
        with open(fpath, "wb") as f:
            f.write(resp.content)
        return fname, True

    return fname, False


# ---------------------------------------------------------------------------
# STEP 4 — COLLECT ALL FILINGS
# ---------------------------------------------------------------------------
def collect_filings(companies: dict, use_cached: bool, log) -> list[dict]:
    """
    Main collection loop. For each company:
        1. Query EDGAR for original 10-K and 10-Q filings
        2. Download up to TARGET_10K / TARGET_10Q most recent
        3. Track results for the summary spreadsheet

    Returns list of per-company result dicts.
    """
    results = []
    total = len(companies)

    for idx, (cik, info) in enumerate(companies.items()):
        ticker = sanitize_ticker(info["ticker"])
        name   = info.get("name", "")
        log.info(f"[{idx+1}/{total}] {ticker} ({name[:45]})")

        ticker_dir = FILINGS_DIR / ticker
        tenk_dir   = ticker_dir / "10-K"
        tenq_dir   = ticker_dir / "10-Q"

        tenk_downloaded = 0
        tenq_downloaded = 0
        tenk_dates = []
        tenq_dates = []

        # --- 10-K filings ---
        filings_10k = get_original_filings(cik, "10-K", TARGET_10K)
        for fil in filings_10k:
            fname, ok = download_filing(fil, ticker, tenk_dir, use_cached)
            if ok:
                tenk_downloaded += 1
                tenk_dates.append(fil["date"])
                log.info(f"    ✓ {fname}")
            else:
                log.warning(f"    ✗ 10-K download failed: {fil['url']}")

        # --- 10-Q filings ---
        filings_10q = get_original_filings(cik, "10-Q", TARGET_10Q)
        for fil in filings_10q:
            fname, ok = download_filing(fil, ticker, tenq_dir, use_cached)
            if ok:
                tenq_downloaded += 1
                tenq_dates.append(fil["date"])
                log.info(f"    ✓ {fname}")
            else:
                log.warning(f"    ✗ 10-Q download failed: {fil['url']}")

        # Coverage status
        if tenk_downloaded >= TARGET_10K and tenq_downloaded >= TARGET_10Q:
            status = "Complete"
        elif tenk_downloaded >= 1 or tenq_downloaded >= 1:
            status = "Partial"
        else:
            status = "No filings"

        results.append({
            "ticker":     ticker,
            "name":       name,
            "cik":        cik,
            "sic":        info.get("sic", ""),
            "sector":     info.get("sector", ""),
            "exchange":   info.get("exchange", ""),
            "10k_count":  tenk_downloaded,
            "10k_dates":  sorted(tenk_dates),
            "10q_count":  tenq_downloaded,
            "10q_dates":  sorted(tenq_dates),
            "status":     status,
        })

    return results


# ---------------------------------------------------------------------------
# STEP 5 — EXCEL SUMMARY
# ---------------------------------------------------------------------------
def generate_excel(results: list[dict], log):
    """Generate the summary spreadsheet with Filing Summary + Statistics."""
    wb = Workbook()

    # === Sheet 1: Filing Summary ==========================================
    ws = wb.active
    ws.title = "Filing Summary"

    hdr_font  = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )
    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")
    amber = PatternFill("solid", fgColor="FFEB9C")

    headers = [
        "Ticker", "Company Name", "CIK", "SIC", "Sub-Sector", "Exchange",
        "10-K Count", "10-K Dates",
        "10-Q Count", "10-Q Dates",
        "Total Filings", "Status",
    ]
    widths = [10, 35, 12, 8, 30, 14, 10, 28, 10, 36, 12, 14]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(c)].width = w

    for r, comp in enumerate(sorted(results, key=lambda x: x["ticker"]), 2):
        total = comp["10k_count"] + comp["10q_count"]
        vals = [
            comp["ticker"],
            comp["name"],
            comp["cik"],
            comp["sic"],
            comp["sector"],
            comp["exchange"],
            comp["10k_count"],
            "; ".join(comp["10k_dates"]),
            comp["10q_count"],
            "; ".join(comp["10q_dates"]),
            total,
            comp["status"],
        ]

        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.font   = Font(name="Arial", size=10)

            if c == 12:  # Status column
                if val == "Complete":
                    cell.fill = green
                elif val == "Partial":
                    cell.fill = amber
                else:
                    cell.fill = red

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{len(results)+1}"
    )

    # === Sheet 2: Statistics ==============================================
    ws2 = wb.create_sheet("Statistics")
    ws2["A1"] = "DDDS Filing Collection — Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")

    complete = sum(1 for r in results if r["status"] == "Complete")
    partial  = sum(1 for r in results if r["status"] == "Partial")
    none_    = sum(1 for r in results if r["status"] == "No filings")
    total_filings = sum(r["10k_count"] + r["10q_count"] for r in results)

    sector_counts = defaultdict(int)
    for r in results:
        sector_counts[r["sector"]] += 1

    exchange_counts = defaultdict(int)
    for r in results:
        exchange_counts[r["exchange"]] += 1

    rows = [
        ("", ""),
        ("Anchor Date", "2025-10-01"),
        ("Collection Window", f"{WINDOW_START}  to  {WINDOW_END}"),
        ("Target per company", f"{TARGET_10K} × 10-K,  {TARGET_10Q} × 10-Q"),
        ("SIC Range", f"{SIC_START}–{SIC_END}"),
        ("Exclusions",
         "Pure-OTC listings, foreign private issuers, "
         "amendments (10-K/A, 10-Q/A)"),
        ("", ""),
        ("Total Companies", len(results)),
        ("Fully covered (Complete)", complete),
        ("Partial coverage", partial),
        ("No filings in window", none_),
        ("Total filings downloaded", total_filings),
        ("", ""),
        ("BY SUB-SECTOR", "Count"),
    ]
    for sect in sorted(sector_counts):
        rows.append((f"  {sect}", sector_counts[sect]))

    rows.append(("", ""))
    rows.append(("BY EXCHANGE", "Count"))
    for exch in sorted(exchange_counts):
        rows.append((f"  {exch or 'Unknown'}", exchange_counts[exch]))

    for i, (label, val) in enumerate(rows, 2):
        ws2[f"A{i}"] = label
        ws2[f"B{i}"] = val
        if label and label == label.upper() and label.strip():
            ws2[f"A{i}"].font = Font(bold=True, name="Arial")

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 45

    wb.save(str(EXCEL_OUTPUT))
    log.info(f"Excel saved to {EXCEL_OUTPUT}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="DDDS - SEC Filing Collection Pipeline"
    )
    parser.add_argument(
        "--use-cached", action="store_true",
        help="Reuse cached universe JSON and skip re-downloading "
             "filings that already exist on disk."
    )
    args = parser.parse_args()
    use_cached = args.use_cached

    log = setup_logging()
    start_time = time.time()

    log.info("=" * 60)
    log.info("DDDS SEC Filing Collection Pipeline")
    log.info(f"Window:  {WINDOW_START} → {WINDOW_END}")
    log.info(f"Targets: {TARGET_10K} × 10-K  |  {TARGET_10Q} × 10-Q")
    log.info(f"SIC:     {SIC_START}–{SIC_END}")
    log.info(f"Cached:  {use_cached}")
    log.info("=" * 60)

    # --- Step 1: Build / load universe ---
    log.info("\n[1/3] Building company universe...")
    companies = build_company_universe(log)
    if not companies:
        log.error("No companies found. Exiting.")
        return
    log.info(f"Universe size: {len(companies)} companies\n")

    # --- Step 2: Collect filings ---
    log.info("[2/3] Collecting filings...")
    results = collect_filings(companies, use_cached, log)

    # --- Step 3: Generate Excel ---
    log.info("\n[3/3] Generating summary spreadsheet...")
    generate_excel(results, log)

    # --- Final summary ---
    elapsed = time.time() - start_time
    complete = sum(1 for r in results if r["status"] == "Complete")
    total_filings = sum(r["10k_count"] + r["10q_count"] for r in results)

    log.info("\n" + "=" * 60)
    log.info(f"Done. {len(results)} companies processed "
             f"in {elapsed/60:.1f} minutes.")
    log.info(f"  Complete coverage:  {complete}/{len(results)}")
    log.info(f"  Total filings:     {total_filings}")
    log.info(f"  HTTP requests:     {_request_count}")
    log.info(f"  Filings dir:       {FILINGS_DIR}/")
    log.info(f"  Excel:             {EXCEL_OUTPUT}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()