"""
DDDS - Backtesting Framework (v5 — Annual, Li (2008) Aligned)
===============================================================
Tests the core hypothesis: disclosure degradation detected by DDDS
predicts future earnings persistence problems.

Methodology follows Li (2008):
    "Annual report readability, current earnings, and earnings persistence"
    Journal of Accounting and Economics, 45(2-3), pp.221-247.

Regression model
----------------
    Earn(t+1) = α + β1·Earn(t) + β2·LOF(t) + β3·LOF(t)·Earn(t)
                  + β4·ln(Assets) + Year_FE + ε

Where:
    Earn(t)     : current-year scaled operating earnings (OI / Assets)
    Earn(t+1)   : next-year scaled operating earnings
    LOF(t)      : Linguistic Obfuscation Factor (continuous, company-year mean)
    LOF×Earn(t) : interaction — KEY TEST VARIABLE
    ln(Assets)  : size control
    Year_FE     : fiscal year fixed effects

Key test: β3 (LOF × Earnings interaction)
    Profitable:  β3 < 0 → obfuscation hides transitory profits (Li's main finding)
    Loss-making: β3 > 0 → obfuscation hides persistent losses  (predicted but NOT
                           found by Li; a significant result here is a novel finding)

LOF scores are aggregated to company-fiscal-year level (mean across all
filings within that fiscal year: 10-K and 10-Q combined).

All companies are US Industrials (SIC 3400-3599), so sector demeaning
is unnecessary.

Data sources:
    - LOF scores:      DDDS pipeline output (score CSVs from run 6)
    - Financial data:  SEC EDGAR XBRL (primary) + yfinance (gap-fill)

Usage
-----
    python run_11_-_backtesting.py
    python run_11_-_backtesting.py --scores-dir data/analyst_outputs
    python run_11_-_backtesting.py --skip-fetch
"""

import argparse
import os
import re
import time
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from dotenv import load_dotenv
from scipy import stats as scipy_stats

load_dotenv()
warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
SCORES_DIR     = os.environ.get("SCORES_DIR", "data/analyst_outputs")
UNIVERSE_FILE  = "company_filing_summary.xlsx"
DATA_CACHE     = "data/backtest/annual_financials_v5.csv"
OUTPUT_DIR     = "data/backtest"
OUTPUT_XLSX    = "data/backtest/backtest_results.xlsx"

EDGAR_HEADERS  = {"User-Agent": os.environ.get(
    "EDGAR_USER_AGENT", "DDDS Research brad@lancaster.ac.uk"
)}
EDGAR_FACTS    = "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
EDGAR_DELAY    = 0.12

MIN_SENTENCES  = 10
WINSORIZE_PCT  = 0.01

REGRESSION_SPECS = [
    ("Vague",    "vague_both",    "Vague (delta-baseline) — MD&A + Risk Factors"),
    ("Complex",  "complex_both",  "Complex (delta-baseline) — MD&A + Risk Factors"),
    ("Combined", "combined_both", "Combined (Vague + Complex)"),
]


# ===================================================================
# STEP 1 — PARSE DDDS SCORE CSVs
# ===================================================================

def _parse_filename(filename: str) -> dict | None:
    m = re.match(
        r"^(.+?)_(10-[KQ])_(\d{4}-\d{2}-\d{2})_(.+?)_ddds_scores\.csv$",
        filename,
    )
    if not m:
        return None
    return {
        "ticker":      m.group(1),
        "filing_type": m.group(2),
        "filing_date": m.group(3),
        "accession":   m.group(4),
    }


def _parse_score_csv(filepath: Path) -> dict | None:
    meta = _parse_filename(filepath.name)
    if not meta:
        return None

    try:
        lines = filepath.read_text(encoding="utf-8-sig").splitlines()
    except Exception:
        return None

    header_idx = None
    for i, line in enumerate(lines):
        if line.strip().startswith("Section,Sentences") or \
           line.strip().startswith("Section\tSentences"):
            header_idx = i
            break
    if header_idx is None:
        return None

    delim = "," if "," in lines[header_idx] else "\t"
    sections = {}
    for i in range(header_idx + 1, min(header_idx + 10, len(lines))):
        line = lines[i].strip()
        if not line or line.startswith("SENTENCE"):
            break
        parts = line.split(delim)
        if len(parts) >= 6:
            name = parts[0].strip()
            try:
                sents = int(parts[1].strip())
                vague = float(parts[3].strip())
                compl = float(parts[5].strip())
            except (ValueError, IndexError):
                continue
            sections[name] = {"sentences": sents, "vague": vague, "complex": compl}

    mda  = sections.get("MD&A")
    risk = sections.get("Risk Factors")
    row  = {**meta}

    row["vague_mda"]   = mda["vague"]   if mda  and mda["sentences"]  >= MIN_SENTENCES else np.nan
    row["complex_mda"] = mda["complex"] if mda  and mda["sentences"]  >= MIN_SENTENCES else np.nan
    row["vague_risk"]  = risk["vague"]  if risk and risk["sentences"] >= MIN_SENTENCES else np.nan
    row["complex_risk"]= risk["complex"]if risk and risk["sentences"] >= MIN_SENTENCES else np.nan

    return row


def load_scores(scores_dir: str) -> pd.DataFrame:
    path = Path(scores_dir)
    if not path.exists():
        print(f"  [!] Scores directory not found: {scores_dir}")
        return pd.DataFrame()

    files = list(path.glob("*_ddds_scores.csv"))
    print(f"  Found {len(files)} score files")

    rows, errors = [], 0
    for f in files:
        r = _parse_score_csv(f)
        if r:
            rows.append(r)
        else:
            errors += 1
    if errors:
        print(f"  [!] {errors} files failed to parse")

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    df["vague_both"]    = df[["vague_mda", "vague_risk"]].mean(axis=1)
    df["complex_both"]  = df[["complex_mda", "complex_risk"]].mean(axis=1)
    df["combined_both"] = df["vague_both"] + df["complex_both"]

    df["filing_date_dt"] = pd.to_datetime(df["filing_date"])

    print(f"  Parsed {len(df)} filings across {df['ticker'].nunique()} companies")
    print(f"  With both sections >={MIN_SENTENCES} sentences: {df['vague_both'].notna().sum()}")
    return df


# ===================================================================
# STEP 2 — MAP FILINGS TO FISCAL YEARS & AGGREGATE
# ===================================================================

def _filing_to_fiscal_year(dt: pd.Timestamp, ftype: str) -> int:
    """
    Map a filing date to the fiscal year it reports on.
    Most SIC 3400-3599 firms have Dec 31 fiscal year ends.

    10-K filed Jan-Jun  -> reports on prior FY (filed ~60-90 days after Dec 31)
    10-K filed Jul-Dec  -> reports on current FY (non-Dec FY end)
    10-Q filed Jan-Mar  -> reports on prior FY quarter (Q3/Q4)
    10-Q filed Apr-Dec  -> reports on current FY quarter (Q1-Q3)
    """
    if ftype == "10-K":
        return dt.year - 1 if dt.month <= 6 else dt.year
    else:
        return dt.year - 1 if dt.month <= 3 else dt.year


def aggregate_to_company_year(scores_df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate per-filing LOF scores to one observation per company-year.
    Uses mean across all filings within a fiscal year.
    """
    df = scores_df.copy()
    df["fiscal_year"] = df.apply(
        lambda r: _filing_to_fiscal_year(r["filing_date_dt"], r["filing_type"]),
        axis=1,
    )

    lof_cols = ["vague_both", "complex_both", "combined_both"]
    agg = df.groupby(["ticker", "fiscal_year"]).agg(
        n_filings=("filing_type", "count"),
        **{col: (col, "mean") for col in lof_cols},
    ).reset_index()

    print(f"  Aggregated to {len(agg)} company-year observations")
    print(f"  Fiscal years covered: {sorted(agg['fiscal_year'].unique())}")
    print(f"  Companies: {agg['ticker'].nunique()}")
    return agg


# ===================================================================
# STEP 3 — FETCH ANNUAL FINANCIAL DATA (EDGAR + YFINANCE)
# ===================================================================

def _edgar_annual(cik: str) -> dict[int, dict]:
    """
    Fetch annual Operating Income and Total Assets from EDGAR XBRL.
    Returns {year: {oi: float, assets: float}}.
    """
    url = EDGAR_FACTS.format(cik=str(cik).zfill(10))
    time.sleep(EDGAR_DELAY)

    try:
        resp = requests.get(url, headers=EDGAR_HEADERS, timeout=15)
        if resp.status_code != 200:
            return {}
        facts = resp.json()
    except Exception:
        return {}

    gaap = facts.get("facts", {}).get("us-gaap", {})

    # Operating income: flow item, must have duration ~350-380 days
    oi = {}
    for concept in ["OperatingIncomeLoss", "OperatingIncome"]:
        entries = gaap.get(concept, {}).get("units", {}).get("USD", [])
        for e in entries:
            if e.get("form", "") not in ("10-K", "10-K/A"):
                continue
            start, end, val = e.get("start", ""), e.get("end", ""), e.get("val")
            if not start or not end or val is None:
                continue
            try:
                days = (pd.Timestamp(end) - pd.Timestamp(start)).days
            except Exception:
                continue
            if 300 <= days <= 400:
                year = int(end[:4])
                if year not in oi:
                    oi[year] = float(val)
        if oi:
            break

    # Total assets: instant item (no duration)
    assets = {}
    for e in gaap.get("Assets", {}).get("units", {}).get("USD", []):
        if e.get("form", "") not in ("10-K", "10-K/A"):
            continue
        end, val = e.get("end", ""), e.get("val")
        if not end or val is None:
            continue
        year = int(end[:4])
        if year not in assets:
            assets[year] = float(val)

    result = {}
    for year in set(oi) & set(assets):
        if assets[year] and assets[year] != 0:
            result[year] = {"oi": oi[year], "assets": assets[year]}
    return result


def _yfinance_annual(ticker: str) -> dict[int, dict]:
    """
    Fetch annual Operating Income and Total Assets from yfinance.
    Returns {year: {oi: float, assets: float}}.
    """
    try:
        import yfinance as yf
    except ImportError:
        return {}

    try:
        stock = yf.Ticker(ticker)
        inc = stock.income_stmt
        bal = stock.balance_sheet
    except Exception:
        return {}

    if inc is None or bal is None or inc.empty or bal.empty:
        return {}

    def _find(df, names):
        for n in names:
            if n in df.index:
                return df.loc[n]
        return None

    oi_row = _find(inc, [
        "Operating Income", "OperatingIncome",
        "Operating Income Loss", "OperatingIncomeLoss",
    ])
    assets_row = _find(bal, ["Total Assets", "TotalAssets"])

    if oi_row is None or assets_row is None:
        return {}

    result = {}
    for dt in assets_row.index:
        try:
            a = float(assets_row[dt])
            if pd.isna(a) or a == 0:
                continue
        except (TypeError, ValueError):
            continue

        try:
            o = float(oi_row[dt]) if dt in oi_row.index and pd.notna(oi_row[dt]) else None
        except (TypeError, ValueError):
            o = None
        if o is None:
            continue

        year = dt.year
        result[year] = {"oi": o, "assets": a}

    return result


def fetch_all_financials(
    tickers: list[str],
    ticker_to_cik: dict[str, str],
    skip_fetch: bool,
) -> pd.DataFrame:
    """
    Fetch annual financials for all tickers.
    Strategy: EDGAR XBRL first, yfinance fills any gaps.
    Returns DataFrame: ticker, year, soe, ln_assets.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    cache = Path(DATA_CACHE)

    if skip_fetch and cache.exists():
        print(f"  Loading cached data from {cache}")
        return pd.read_csv(cache)

    print(f"  Fetching annual data for {len(tickers)} companies...")
    print(f"  Strategy: EDGAR XBRL (primary) + yfinance (gap-fill)\n")

    all_rows = []
    edgar_count, yf_count, fail_count = 0, 0, 0

    for i, ticker in enumerate(tickers):
        print(f"  [{i+1}/{len(tickers)}] {ticker}", end=" ... ")

        cik = ticker_to_cik.get(ticker)
        edgar_data = _edgar_annual(cik) if cik else {}
        yf_data = _yfinance_annual(ticker)

        # Merge: EDGAR priority, yfinance fills gaps
        all_years = set(edgar_data.keys()) | set(yf_data.keys())

        if not all_years:
            print("no data")
            fail_count += 1
            continue

        company_rows = []
        for year in sorted(all_years):
            if year in edgar_data:
                d, src = edgar_data[year], "edgar"
            elif year in yf_data:
                d, src = yf_data[year], "yfinance"
            else:
                continue

            if d["assets"] == 0:
                continue

            company_rows.append({
                "ticker":    ticker,
                "year":      year,
                "soe":       d["oi"] / d["assets"],
                "ln_assets": np.log(abs(d["assets"])),
                "source":    src,
            })

        all_rows.extend(company_rows)
        e_n = sum(1 for r in company_rows if r["source"] == "edgar")
        y_n = sum(1 for r in company_rows if r["source"] == "yfinance")

        parts = [f"edgar:{e_n}"] if e_n else []
        if y_n:
            parts.append(f"yf:{y_n}")
        print(f"{len(company_rows)} years ({' '.join(parts)})")

        if e_n:
            edgar_count += 1
        if y_n:
            yf_count += 1

    df = pd.DataFrame(all_rows)
    if not df.empty:
        df.to_csv(cache, index=False)
        print(f"\n  Cached to {cache}")
        print(f"  EDGAR: {edgar_count} companies | yfinance gap-fill: {yf_count} | failed: {fail_count}")
        print(f"  Total company-years: {len(df)}")
    return df


# ===================================================================
# STEP 4 — BUILD REGRESSION DATASET
# ===================================================================

def build_regression_dataset(
    lof_df: pd.DataFrame,
    fin_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Merge company-year LOF with annual financials.
    Creates Earn(t) and Earn(t+1) for each company-year.
    """
    fin_lookup = {}
    for _, row in fin_df.iterrows():
        fin_lookup[(row["ticker"], int(row["year"]))] = {
            "soe":       row["soe"],
            "ln_assets": row["ln_assets"],
        }

    rows = []
    for _, obs in lof_df.iterrows():
        ticker = obs["ticker"]
        fy     = int(obs["fiscal_year"])

        t_data  = fin_lookup.get((ticker, fy))
        t1_data = fin_lookup.get((ticker, fy + 1))

        if t_data is None:
            continue

        row = obs.to_dict()
        row["earnings_t"]  = t_data["soe"]
        row["ln_assets"]   = t_data["ln_assets"]
        row["earnings_t1"] = t1_data["soe"] if t1_data else np.nan
        rows.append(row)

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    for col in ["earnings_t", "earnings_t1"]:
        valid = df[col].dropna()
        if len(valid) > 20:
            lo = valid.quantile(WINSORIZE_PCT)
            hi = valid.quantile(1 - WINSORIZE_PCT)
            df[col] = df[col].clip(lower=lo, upper=hi)

    n_complete = df.dropna(subset=["earnings_t", "earnings_t1", "ln_assets"]).shape[0]

    print(f"\n  Regression dataset: {len(df)} company-years")
    print(f"  With Earn(t):      {df['earnings_t'].notna().sum()}")
    print(f"  With Earn(t+1):    {df['earnings_t1'].notna().sum()}")
    print(f"  Complete cases:    {n_complete}")
    return df


# ===================================================================
# STEP 5 — REGRESSION ENGINE
# ===================================================================

def run_regression(
    df: pd.DataFrame,
    lof_col: str,
    subsample: str = "full",
) -> dict:
    """
    Li (2008) earnings persistence regression with cluster-robust SEs.

    Earn(t+1) = a + b1*Earn(t) + b2*LOF + b3*LOF*Earn(t)
                  + b4*ln(Assets) + Year_FE + e
    """
    required = ["earnings_t", "earnings_t1", lof_col, "ln_assets", "fiscal_year"]
    sub = df.dropna(subset=required).copy()

    if subsample == "profitable":
        sub = sub[sub["earnings_t"] > 0]
    elif subsample == "loss":
        sub = sub[sub["earnings_t"] < 0]

    n = len(sub)
    n_firms = sub["ticker"].nunique() if n > 0 else 0

    if n < 15:
        return {
            "subsample": subsample, "n": n, "n_firms": n_firms,
            "error": f"Insufficient observations ({n})",
        }

    y     = sub["earnings_t1"].values
    earn  = sub["earnings_t"].values
    lof   = sub[lof_col].values
    inter = lof * earn
    lna   = sub["ln_assets"].values

    core = [np.ones(n), earn, lof, inter, lna]
    core_labels = ["a", "b1_Earn", "b2_LOF", "b3_LOFxEarn", "b4_lnAssets"]

    # Year FE (only if 3+ years in subsample)
    years = sub["fiscal_year"].values
    unique_years = sorted(np.unique(years))
    fe_cols, fe_labels = [], []
    if len(unique_years) >= 3:
        for yr in unique_years[1:]:
            d = (years == yr).astype(float)
            if 0 < d.sum() < n:
                fe_cols.append(d)
                fe_labels.append(f"FE_{yr}")

    X = np.column_stack(core + fe_cols) if fe_cols else np.column_stack(core)
    labels = core_labels + fe_labels
    k = X.shape[1]

    # Rank check — drop FE columns if needed
    rank = np.linalg.matrix_rank(X)
    while rank < k and fe_labels:
        fe_cols.pop()
        fe_labels.pop()
        X = np.column_stack(core + fe_cols) if fe_cols else np.column_stack(core)
        labels = core_labels + fe_labels
        k = X.shape[1]
        rank = np.linalg.matrix_rank(X)

    # OLS
    try:
        coeffs, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
    except np.linalg.LinAlgError as e:
        return {"subsample": subsample, "n": n, "n_firms": n_firms, "error": str(e)}

    resid = y - X @ coeffs
    ss_res = np.sum(resid ** 2)
    ss_tot = np.sum((y - y.mean()) ** 2)
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
    adj_r2 = 1 - (1 - r2) * (n - 1) / (n - k) if n > k else r2

    # Cluster-robust SEs
    tickers_arr = sub["ticker"].values
    unique_tickers = np.unique(tickers_arr)
    n_clusters = len(unique_tickers)

    try:
        XtX_inv = np.linalg.inv(X.T @ X)
    except np.linalg.LinAlgError:
        dof_fallback = n - k
        mse = ss_res / dof_fallback if dof_fallback > 0 else np.nan
        cov = mse * np.linalg.pinv(X.T @ X)
        se = np.sqrt(np.maximum(np.diag(cov), 0.0))
        n_clusters = 0
        XtX_inv = None

    if XtX_inv is not None:
        meat = np.zeros((k, k))
        for t in unique_tickers:
            mask = tickers_arr == t
            X_g = X[mask]
            e_g = resid[mask]
            score = X_g.T @ e_g
            meat += np.outer(score, score)

        corr = (n_clusters / (n_clusters - 1)) * ((n - 1) / (n - k)) \
               if n_clusters > 1 else 1.0
        cov = corr * XtX_inv @ meat @ XtX_inv
        se = np.sqrt(np.maximum(np.diag(cov), 0.0))

    dof = min(n_clusters - 1, n - k) if n_clusters > 1 else n - k

    with np.errstate(divide="ignore", invalid="ignore"):
        t_stats = np.where(se > 0, coeffs / se, np.nan)

    p_vals = [
        2 * (1 - scipy_stats.t.cdf(abs(t), df=dof)) if not np.isnan(t) else np.nan
        for t in t_stats
    ]

    result = {
        "subsample":  subsample,
        "n":          n,
        "n_firms":    n_firms,
        "n_clusters": n_clusters,
        "r2":         round(r2, 4),
        "adj_r2":     round(adj_r2, 4),
        "dof":        dof,
        "n_year_fe":  len(fe_labels),
    }

    for i, label in enumerate(core_labels):
        result[f"coeff_{label}"] = round(float(coeffs[i]), 6)
        result[f"se_{label}"]    = round(float(se[i]), 6)
        result[f"t_{label}"]     = round(float(t_stats[i]), 4) if not np.isnan(t_stats[i]) else None
        result[f"p_{label}"]     = round(float(p_vals[i]), 6) if not np.isnan(p_vals[i]) else None

    b3 = float(coeffs[3])
    p3 = float(p_vals[3]) if not np.isnan(p_vals[3]) else 1.0
    sig = p3 < 0.10

    if subsample == "loss":
        result["expected_sign"] = "b3 > 0"
        result["hypothesis_supported"] = sig and b3 > 0
    else:
        result["expected_sign"] = "b3 < 0"
        result["hypothesis_supported"] = sig and b3 < 0

    return result


# ===================================================================
# STEP 6 — OUTPUT
# ===================================================================

def print_summary(all_results: dict):
    print(f"\n{'='*92}")
    print(f"  DDDS BACKTESTING — Li (2008) Annual Earnings Persistence")
    print(f"  Earn(t+1) = a + b1*Earn(t) + b2*LOF + b3*LOF*Earn(t) + b4*ln(Assets) + Year_FE")
    print(f"  Cluster-robust SEs by ticker | DoF = min(G-1, N-k)")
    print(f"{'='*92}")
    print(f"\n  {'Spec':<18} {'Sub':<12} {'N':>5} {'Firms':>5} {'Expected':>8} "
          f"{'b3':>10} {'t(b3)':>8} {'p(b3)':>8} {'Adj R2':>7} {'H0':>4}")
    print(f"  {'-'*87}")

    for _, lof_col, description in REGRESSION_SPECS:
        results = all_results.get(lof_col, [])
        label = description.split(" — ")[0].strip() if " — " in description else description

        for r in results:
            if "error" in r:
                print(f"  {label:<18} {r['subsample']:<12} {r['n']:>5} "
                      f"{r.get('n_firms',''):>5} {'':>8} {'ERROR':>10}")
                label = ""
                continue

            h0 = "Y" if r.get("hypothesis_supported") else "N"
            b3 = f"{r['coeff_b3_LOFxEarn']:.6f}"
            t3 = f"{r['t_b3_LOFxEarn']:.3f}" if r.get("t_b3_LOFxEarn") is not None else "N/A"
            p3 = f"{r['p_b3_LOFxEarn']:.4f}" if r.get("p_b3_LOFxEarn") is not None else "N/A"

            print(f"  {label:<18} {r['subsample']:<12} {r['n']:>5} {r['n_firms']:>5} "
                  f"{r.get('expected_sign',''):>8} {b3:>10} {t3:>8} {p3:>8} "
                  f"{r['adj_r2']:>7.4f} {h0:>4}")
            label = ""

        print(f"  {'-'*87}")

    print(f"\n  Li (2008) predictions:")
    print(f"    Profitable: b3 < 0 -> more complex reports hide transitory good news")
    print(f"    Loss:       b3 > 0 -> more complex reports hide persistent bad news")
    print(f"    (Li found significance only for profitable firms; loss result is novel)")
    print(f"{'='*92}\n")


def write_excel(all_results: dict, reg_df: pd.DataFrame, output: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    wb.remove(wb.active)

    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    body     = Font(name="Arial", size=10)
    title_f  = Font(name="Arial", bold=True, size=13)
    note_f   = Font(name="Arial", size=9, italic=True)
    sig_fill = PatternFill("solid", fgColor="C6EFCE")
    no_fill  = PatternFill("solid", fgColor="FFC7CE")

    core = ["a", "b1_Earn", "b2_LOF", "b3_LOFxEarn", "b4_lnAssets"]

    for sheet_name, lof_col, desc in REGRESSION_SPECS:
        results = all_results.get(lof_col, [])
        ws = wb.create_sheet(title=sheet_name)

        ws["A1"] = "DDDS Backtesting — Li (2008) Annual Earnings Persistence"
        ws["A1"].font = title_f
        ws["A2"] = f"LOF: {desc}"
        ws["A2"].font = Font(name="Arial", size=11, italic=True)
        ws["A3"] = "Earn(t+1) = a + b1*Earn(t) + b2*LOF + b3*LOF*Earn(t) + b4*ln(Assets) + Year FE"
        ws["A3"].font = body

        row = 5
        headers = ["Subsample", "N", "Firms", "Clusters", "DoF", "Adj R2", "Year FEs", "Expected"]
        for c in core:
            headers += [f"Coeff({c})", f"SE({c})", f"t({c})", f"p({c})"]
        headers.append("H0 Supported")

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font, cell.fill = hdr_font, hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for r in results:
            row += 1
            if "error" in r:
                ws.cell(row=row, column=1, value=r["subsample"]).font = body
                ws.cell(row=row, column=2, value=r["n"]).font = body
                ws.cell(row=row, column=6, value=f"ERROR: {r['error']}").font = body
                continue

            vals = [r["subsample"], r["n"], r["n_firms"], r["n_clusters"],
                    r["dof"], r["adj_r2"], r["n_year_fe"], r.get("expected_sign", "")]
            for c in core:
                vals += [r.get(f"coeff_{c}"), r.get(f"se_{c}"),
                         r.get(f"t_{c}"), r.get(f"p_{c}")]
            vals.append("Yes" if r.get("hypothesis_supported") else "No")

            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = body
                if isinstance(v, float):
                    cell.number_format = "0.0000"

            h0_cell = ws.cell(row=row, column=len(headers))
            h0_cell.fill = sig_fill if r.get("hypothesis_supported") else no_fill

        row += 2
        notes = [
            f"Cluster-robust SEs by ticker. DoF = min(G-1, N-k). "
            f"Earnings winsorized at {WINSORIZE_PCT:.0%}/{1-WINSORIZE_PCT:.0%}. "
            f"Sections < {MIN_SENTENCES} sentences excluded. "
            f"LOF aggregated to company-year mean. All firms SIC 3400-3599.",

            "H0: Profitable b3<0 (Li's main finding); Loss b3>0 (predicted by Li, "
            "not confirmed in original paper — significant result here is novel).",

            "Li, F. (2008). Annual report readability, current earnings, and earnings "
            "persistence. JAE 45(2-3), 221-247.",

            "Data: EDGAR XBRL (primary) + yfinance (gap-fill). "
            "LOF = DDDS Linguistic Obfuscation Factor from fine-tuned FinBERT classifiers.",
        ]
        for note in notes:
            ws.cell(row=row, column=1, value=note).font = note_f
            row += 1

        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=5, column=ci).column_letter].width = 14

    # Data sheet
    dws = wb.create_sheet(title="Data")
    export = ["ticker", "fiscal_year", "n_filings",
              "vague_both", "complex_both", "combined_both",
              "earnings_t", "earnings_t1", "ln_assets"]
    cols = [c for c in export if c in reg_df.columns]

    for ci, c in enumerate(cols, 1):
        cell = dws.cell(row=1, column=ci, value=c)
        cell.font, cell.fill = hdr_font, hdr_fill

    for ri, (_, dr) in enumerate(reg_df.iterrows(), 2):
        for ci, c in enumerate(cols, 1):
            v = dr[c]
            dws.cell(row=ri, column=ci, value=None if pd.isna(v) else v).font = body

    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)
    print(f"  Results saved to: {output}")


# ===================================================================
# MAIN
# ===================================================================

def main():
    parser = argparse.ArgumentParser(
        description="DDDS Backtesting — Li (2008) Annual Earnings Persistence"
    )
    parser.add_argument("--scores-dir", default=SCORES_DIR)
    parser.add_argument("--skip-fetch", action="store_true",
                        help="Use cached financial data only")
    parser.add_argument("--output", default=OUTPUT_XLSX)
    args = parser.parse_args()

    print(f"\n[DDDS] Backtesting v5 — Li (2008) Annual Earnings Persistence")
    print(f"  Scores:  {args.scores_dir}")
    print(f"  Output:  {args.output}\n")

    # 1: Load DDDS scores
    print("[1/5] Loading DDDS scores...")
    scores_df = load_scores(args.scores_dir)
    if scores_df.empty:
        print("  [!] No scores loaded.")
        return

    # 2: Aggregate to company-year
    print("\n[2/5] Aggregating LOF scores to company-year...")
    lof_df = aggregate_to_company_year(scores_df)

    # 3: Fetch financials
    print("\n[3/5] Loading universe & fetching annual financials...")
    if not Path(UNIVERSE_FILE).exists():
        print(f"  [!] Universe file not found: {UNIVERSE_FILE}")
        return

    univ = pd.read_excel(UNIVERSE_FILE)
    if "Ticker" not in univ.columns or "CIK" not in univ.columns:
        print("  [!] Universe file missing Ticker/CIK columns")
        return

    univ = univ.dropna(subset=["Ticker", "CIK"])
    ticker_to_cik = {
        row["Ticker"]: str(int(row["CIK"]))
        for _, row in univ.iterrows()
    }
    print(f"  Universe: {len(ticker_to_cik)} companies with CIK")

    tickers = lof_df["ticker"].unique().tolist()
    fin_df = fetch_all_financials(tickers, ticker_to_cik, args.skip_fetch)
    if fin_df.empty:
        print("  [!] No financial data retrieved.")
        return

    # 4: Build regression dataset
    print("\n[4/5] Building regression dataset...")
    reg_df = build_regression_dataset(lof_df, fin_df)
    if reg_df.empty:
        print("  [!] Empty regression dataset.")
        return

    # 5: Run regressions
    print("\n[5/5] Running regressions (3 specs x 3 subsamples)...")
    all_results = {}
    for _, lof_col, _ in REGRESSION_SPECS:
        spec_results = []
        for sub in ["full", "profitable", "loss"]:
            r = run_regression(reg_df, lof_col, sub)
            spec_results.append(r)
        all_results[lof_col] = spec_results

    print_summary(all_results)
    write_excel(all_results, reg_df, args.output)
    print("[Done]\n")


if __name__ == "__main__":
    main()