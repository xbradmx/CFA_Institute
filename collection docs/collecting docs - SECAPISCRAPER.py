"""
DDDS Data Collector
===================
Downloads 10-K, 10-Q filings (and amendments) and flags earnings call
transcript availability for companies across Electronics, Materials, and
Energy sectors.

Window:  October 1 2023 → September 30 2024  (12-month lookback)
Anchor:  October 1 2024 — two full quarters of out-of-sample data available.

Directory structure created:
  DDDS/Companies/{TICKER}/10-K/
  DDDS/Companies/{TICKER}/10-Q/
  DDDS/Companies/{TICKER}/Earnings_Transcripts/

Output:
  DDDS/company_filing_summary.xlsx

Usage:
  pip install requests openpyxl tqdm
  python ddds_data_collector.py

IMPORTANT: Update the USER_AGENT below with your real name and email.
SEC will block requests without a proper user-agent.
"""

import os
import json
import time
import re
import requests
import logging
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# ── Configuration ──────────────────────────────────────────────────────────

# !! UPDATE THIS with your real name and email — SEC requires it !!
USER_AGENT = "connorokeeffe07@gmail.com"

ANCHOR_DATE = "2025-10-01"
WINDOW_START = "2024-10-01"
WINDOW_END = "2025-09-30"

BASE_DIR = Path("DDDS/Companies")
OUTPUT_EXCEL = Path("DDDS/company_filing_summary.xlsx")
CACHE_FILE = Path("DDDS/company_universe_cache.json")
LOG_FILE = Path("DDDS/ddds_collector.log")

HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept-Encoding": "gzip, deflate",
}

# SEC rate limit: 10 req/s. We target ~6/s to be safe.
REQUEST_DELAY = 0.15

# Form types to download (includes amendments)
ANNUAL_FORMS = {"10-K", "10-K/A"}
QUARTERLY_FORMS = {"10-Q", "10-Q/A"}
EARNINGS_FORMS = {"8-K", "8-K/A"}
ALL_TARGET_FORMS = ANNUAL_FORMS | QUARTERLY_FORMS | EARNINGS_FORMS

# Coverage universe: SIC 3400–3599 (Fabricated Metals + Industrial/Commercial Machinery)
# 3400–3499: Fabricated Metal Products (except machinery & transport equipment)
# 3500–3599: Industrial & Commercial Machinery & Computer Equipment
SIC_START = 3400
SIC_END = 3599

# OTC exclusion is handled in build_company_universe by checking if ALL
# exchanges are "OTC". Companies with any non-OTC exchange are kept.


# ── Setup ──────────────────────────────────────────────────────────────────

def setup_logging():
    """Set up logging — creates DDDS/ first so FileHandler doesn't crash."""
    os.makedirs("DDDS", exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(str(LOG_FILE)),
            logging.StreamHandler(),
        ],
    )
    return logging.getLogger("ddds")


# ── HTTP Helpers ───────────────────────────────────────────────────────────

# Adaptive delay — starts at REQUEST_DELAY, increases on 429s, decays back slowly
_current_delay = REQUEST_DELAY
_request_count = 0


def rate_limit():
    global _current_delay
    time.sleep(_current_delay)
    # Slowly decay back toward base delay after successful requests
    if _current_delay > REQUEST_DELAY:
        _current_delay = max(REQUEST_DELAY, _current_delay * 0.995)


def safe_get(url, params=None, max_retries=3):
    """GET with retries, adaptive rate limiting, and proper error handling."""
    global _current_delay, _request_count
    for attempt in range(max_retries):
        try:
            rate_limit()
            _request_count += 1
            resp = requests.get(url, headers=HEADERS, params=params, timeout=30)
            if resp.status_code == 200:
                return resp
            elif resp.status_code == 429:
                # Back off globally — double the delay, wait, then continue
                _current_delay = min(_current_delay * 2, 5.0)
                wait = 15 * (attempt + 1)
                logging.warning(
                    f"Rate limited (429). Global delay now {_current_delay:.2f}s. "
                    f"Waiting {wait}s before retry..."
                )
                time.sleep(wait)
                continue
            elif resp.status_code == 404:
                return None  # Don't retry 404s
            else:
                logging.warning(f"HTTP {resp.status_code} for {url}")
                if attempt < max_retries - 1:
                    time.sleep(3)
                    continue
                return None
        except requests.exceptions.RequestException as e:
            logging.warning(f"Request error (attempt {attempt+1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(5)
    return None


def get_request_count():
    return _request_count


def sanitize_ticker(ticker):
    """Clean ticker for use as folder name."""
    return re.sub(r'[^A-Za-z0-9\-.]', '', ticker).upper()


# ── Step 1: Build Company Universe ─────────────────────────────────────────

def build_company_universe(log):
    """
    Build the coverage universe by checking EVERY SEC-registered company's
    SIC code via the submissions API. This is the only reliable method —
    EDGAR's browse-by-SIC search index is incomplete and misses ~60% of
    companies.

    Takes ~19 minutes for ~13k companies at 0.12s per request.
    Results are cached so subsequent runs are instant.

    Filtering:
    - SIC 3400–3599 only
    - Exclude companies whose exchanges are ALL "OTC" (pure OTC)
    - Deduplicate by CIK (first ticker encountered wins — skips warrants,
      preferred shares, dual share classes like GEF-B, POWWP, MOG-B)

    Returns: dict keyed by CIK string.
    """
    if CACHE_FILE.exists():
        log.info(f"Loading cached universe from {CACHE_FILE}")
        with open(CACHE_FILE) as f:
            cached = json.load(f)
        log.info(f"Loaded {len(cached)} companies from cache.")
        return cached

    os.makedirs("DDDS", exist_ok=True)

    # Step 1: Download the full company tickers list
    log.info("Downloading SEC company tickers...")
    resp = safe_get("https://www.sec.gov/files/company_tickers.json")
    if not resp:
        log.error("FATAL: Cannot download company tickers. Check network.")
        return {}

    all_tickers = resp.json()
    total = len(all_tickers)
    log.info(f"Total SEC-registered companies to scan: {total}")
    log.info(f"Estimated time: ~{total * 0.12 / 60:.0f} minutes")

    # Step 2: Check each company's SIC via submissions API
    companies = {}  # Keyed by CIK — naturally deduplicates share classes
    skipped_otc = 0
    skipped_dupes = 0
    errors = 0

    for i, (key, company) in enumerate(tqdm(all_tickers.items(),
                                             desc="Scanning all companies",
                                             total=total)):
        cik = str(company["cik_str"])
        cik_padded = cik.zfill(10)
        ticker = company.get("ticker", "")

        # CIK dedup: if we already have this company, skip
        # (handles GEF/GEF-B, MOG-A/MOG-B, POWW/POWWP, etc.)
        if cik in companies:
            skipped_dupes += 1
            continue

        try:
            url = f"https://data.sec.gov/submissions/CIK{cik_padded}.json"
            rate_limit()
            resp = requests.get(url, headers=HEADERS, timeout=10)
            if resp.status_code != 200:
                errors += 1
                continue

            data = resp.json()
            sic = int(data.get("sic", 0) or 0)

            if SIC_START <= sic <= SIC_END:
                exchanges = data.get("exchanges", [])
                exchanges = [e for e in exchanges if e is not None]

                # OTC exclusion: skip only if EVERY exchange is OTC
                # Companies like NYSE|OTC are fine — they have a major listing
                if exchanges:
                    is_pure_otc = all(e.upper() == "OTC" for e in exchanges)
                    if is_pure_otc:
                        skipped_otc += 1
                        continue

                # Sub-sector label
                if sic < 3500:
                    sub_sector = "Fabricated Metal Products"
                else:
                    sub_sector = "Industrial & Commercial Machinery"

                # Store all exchanges for reference
                exchange_str = " | ".join(exchanges) if exchanges else "Unknown"

                companies[cik] = {
                    "cik": cik,
                    "ticker": ticker,
                    "name": data.get("name", ""),
                    "sic": str(sic),
                    "sic_description": data.get("sicDescription", ""),
                    "sector": sub_sector,
                    "exchange": exchange_str,
                    "category": data.get("category", ""),
                    "website": data.get("website", ""),
                }
                log.info(f"  FOUND: {ticker:8s} | SIC {sic} | {data.get('sicDescription', '')}")

        except Exception:
            errors += 1

        # Progress update every 1000 companies
        if (i + 1) % 1000 == 0:
            log.info(f"  ...scanned {i+1}/{total} | found {len(companies)} | "
                     f"{skipped_dupes} dupes skipped | {skipped_otc} OTC excluded | "
                     f"{errors} errors")

    log.info(f"Scan complete: {len(companies)} unique companies "
             f"(SIC {SIC_START}–{SIC_END}). "
             f"{skipped_dupes} duplicate tickers skipped, "
             f"{skipped_otc} pure-OTC excluded, {errors} errors.")

    # Cache results
    with open(CACHE_FILE, "w") as f:
        json.dump(companies, f, indent=2)
    log.info(f"Cached universe to {CACHE_FILE}")

    return companies


# ── Step 2: Get & Download Filings ─────────────────────────────────────────

def get_filings_for_company(cik, log):
    """
    Fetch filing metadata from the EDGAR submissions API.
    Returns list of filing dicts within our date window.
    """
    cik_padded = str(cik).zfill(10)
    url = f"https://data.sec.gov/submissions/CIK{cik_padded}.json"
    resp = safe_get(url)
    if not resp:
        return []

    try:
        data = resp.json()
    except (json.JSONDecodeError, ValueError):
        log.warning(f"Invalid JSON from submissions API for CIK {cik}")
        return []

    recent = data.get("filings", {}).get("recent", {})
    if not recent:
        return []

    forms = recent.get("form", [])
    dates = recent.get("filingDate", [])
    accessions = recent.get("accessionNumber", [])
    primary_docs = recent.get("primaryDocument", [])

    # Guard against mismatched array lengths
    min_len = min(len(forms), len(dates), len(accessions), len(primary_docs))

    filings = []
    for i in range(min_len):
        form = forms[i]
        if form not in ALL_TARGET_FORMS:
            continue

        filing_date = dates[i]
        if filing_date < WINDOW_START or filing_date > WINDOW_END:
            continue

        accession_formatted = accessions[i]            # "0000320193-24-000081"
        accession_path = accession_formatted.replace("-", "")  # "000032019324000081"
        primary_doc = primary_docs[i]

        # EDGAR archive URL:
        # https://www.sec.gov/Archives/edgar/data/{CIK_PADDED}/{ACCESSION_NO_DASHES}/{DOC}
        filing_url = (
            f"https://www.sec.gov/Archives/edgar/data/"
            f"{cik_padded}/{accession_path}/{primary_doc}"
        )
        index_url = (
            f"https://www.sec.gov/Archives/edgar/data/"
            f"{cik_padded}/{accession_path}/"
        )

        filings.append({
            "form": form,
            "date": filing_date,
            "accession_formatted": accession_formatted,
            "accession_path": accession_path,
            "primary_doc": primary_doc,
            "url": filing_url,
            "index_url": index_url,
        })

    return filings


def download_filing(filing, save_dir, log):
    """Download a filing's primary document. Returns (filepath, success)."""
    os.makedirs(save_dir, exist_ok=True)

    form_clean = filing["form"].replace("/", "-")  # 10-K/A → 10-K-A
    date_str = filing["date"]
    accession = filing["accession_formatted"]

    # Extension from primary doc
    pdoc = filing["primary_doc"].lower()
    if pdoc.endswith(".htm") or pdoc.endswith(".html"):
        ext = ".html"
    elif pdoc.endswith(".txt"):
        ext = ".txt"
    elif pdoc.endswith(".xml"):
        ext = ".xml"
    else:
        ext = ".html"

    filename = f"{form_clean}_{date_str}_{accession}{ext}"
    filepath = os.path.join(save_dir, filename)

    if os.path.exists(filepath):
        return filepath, True

    resp = safe_get(filing["url"])
    if resp and len(resp.content) > 100:
        with open(filepath, "wb") as f:
            f.write(resp.content)
        return filepath, True

    return filepath, False


# ── Step 3: Main Collection Loop ──────────────────────────────────────────

def collect_all_filings(companies, log):
    """Download all filings for every company. Returns summary list."""
    summary = []
    total = len(companies)

    for idx, (cik, info) in enumerate(companies.items()):
        ticker = sanitize_ticker(info["ticker"])
        name = info.get("name", "")
        sector = info.get("sector", "Unknown")
        sic = info.get("sic", "")

        log.info(f"[{idx+1}/{total}] {ticker} ({name})")

        # Create directories
        ticker_dir = BASE_DIR / ticker
        tenk_dir = ticker_dir / "10-K"
        tenq_dir = ticker_dir / "10-Q"
        eightk_dir = ticker_dir / "8-K"
        for d in [tenk_dir, tenq_dir, eightk_dir]:
            os.makedirs(d, exist_ok=True)

        filings = get_filings_for_company(cik, log)
        if not filings:
            log.warning(f"  No filings found for {ticker} (CIK {cik})")

        tenk_count = 0
        tenq_count = 0
        eightk_count = 0
        tenk_dates = []
        tenq_dates = []
        eightk_dates = []

        for filing in filings:
            form = filing["form"]

            if form in ANNUAL_FORMS:
                _, ok = download_filing(filing, str(tenk_dir), log)
                if ok:
                    tenk_count += 1
                    tenk_dates.append(filing["date"])

            elif form in QUARTERLY_FORMS:
                _, ok = download_filing(filing, str(tenq_dir), log)
                if ok:
                    tenq_count += 1
                    tenq_dates.append(filing["date"])

            elif form in EARNINGS_FORMS:
                _, ok = download_filing(filing, str(eightk_dir), log)
                if ok:
                    eightk_count += 1
                    eightk_dates.append(filing["date"])

        has_10k = tenk_count >= 1
        has_10q = tenq_count >= 2

        if has_10k and has_10q:
            completeness = "Complete"
        elif has_10k or tenq_count >= 1:
            completeness = "Partial"
        else:
            completeness = "No filings"

        summary.append({
            "ticker": ticker,
            "name": name,
            "cik": cik,
            "sic": sic,
            "sector": sector,
            "exchange": info.get("exchange", ""),
            "10k_count": tenk_count,
            "10k_dates": "; ".join(sorted(tenk_dates)),
            "10q_count": tenq_count,
            "10q_dates": "; ".join(sorted(tenq_dates)),
            "8k_count": eightk_count,
            "8k_dates": "; ".join(sorted(eightk_dates)),
            "has_10k": has_10k,
            "has_10q": has_10q,
            "total_sec_filings": tenk_count + tenq_count + eightk_count,
            "completeness": completeness,
        })

    return summary


# ── Step 4: Excel Summary ─────────────────────────────────────────────────

def generate_excel(summary, log):
    """Generate the summary spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Filing Summary"

    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    amber = PatternFill("solid", fgColor="FFEB9C")

    headers = [
        "Ticker", "Company Name", "CIK", "SIC", "Sub-Sector", "Exchange",
        "10-K Count", "10-K Dates",
        "10-Q Count", "10-Q Dates",
        "8-K Count", "8-K Dates",
        "Has 10-K?", "Has 10-Qs?",
        "Total SEC Filings", "Completeness",
    ]
    widths = [10, 35, 12, 8, 30, 14, 10, 28, 10, 35, 10, 35, 12, 12, 16, 14]

    for c, (hdr, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w

    for r, comp in enumerate(sorted(summary, key=lambda x: x["ticker"]), 2):
        vals = [
            comp["ticker"],
            comp["name"],
            comp["cik"],
            comp["sic"],
            comp["sector"],
            comp.get("exchange", ""),
            comp["10k_count"],
            comp["10k_dates"],
            comp["10q_count"],
            comp["10q_dates"],
            comp["8k_count"],
            comp["8k_dates"],
            "Yes" if comp["has_10k"] else "No",
            "Yes" if comp["has_10q"] else "No",
            comp["total_sec_filings"],
            comp["completeness"],
        ]

        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=10)

            if c == 13:  # Has 10-K
                cell.fill = green if comp["has_10k"] else red
            elif c == 14:  # Has 10-Qs
                cell.fill = green if comp["has_10q"] else red
            elif c == 16:  # Completeness
                if val == "Complete":
                    cell.fill = green
                elif val == "Partial":
                    cell.fill = amber
                else:
                    cell.fill = red

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(summary)+1}"

    # Summary Statistics sheet
    ws2 = wb.create_sheet("Statistics")
    ws2["A1"] = "DDDS Filing Collection — Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")

    sector_counts = defaultdict(int)
    for s in summary:
        sector_counts[s["sector"]] += 1

    complete = sum(1 for s in summary if s["completeness"] == "Complete")
    partial = sum(1 for s in summary if s["completeness"] == "Partial")
    none_ = sum(1 for s in summary if s["completeness"] == "No filings")
    total_8ks = sum(s["8k_count"] for s in summary)

    rows = [
        ("", ""),
        ("Anchor Date (as-if run date)", ANCHOR_DATE),
        ("Filing Window", f"{WINDOW_START}  to  {WINDOW_END}"),
        ("SIC Range", f"{SIC_START}–{SIC_END}"),
        ("OTC Excluded", "Yes (pure-OTC only)"),
        ("Total Companies", len(summary)),
        ("", ""),
        ("BY SUB-SECTOR", "Count"),
    ]
    for sect in sorted(sector_counts):
        rows.append((f"  {sect}", sector_counts[sect]))

    # Exchange breakdown
    exchange_counts = defaultdict(int)
    for s in summary:
        exchange_counts[s.get("exchange", "Unknown")] += 1
    rows.append(("", ""))
    rows.append(("BY EXCHANGE", "Count"))
    for exch in sorted(exchange_counts):
        rows.append((f"  {exch or 'Unknown'}", exchange_counts[exch]))

    rows += [
        ("", ""),
        ("FILING COVERAGE", ""),
        ("Complete (10-K + 2+ 10-Q)", complete),
        ("Partial", partial),
        ("No filings in window", none_),
        ("", ""),
        ("8-K FILINGS", ""),
        ("Total 8-Ks downloaded", total_8ks),
        ("", ""),
        ("NOTE", "8-K filings contain press releases and event"),
        ("", "disclosures. Earnings call transcripts must be"),
        ("", "sourced separately from a transcript provider."),
    ]

    for i, (label, val) in enumerate(rows, 2):
        ws2[f"A{i}"] = label
        ws2[f"B{i}"] = val
        if label and label == label.upper() and label.strip():
            ws2[f"A{i}"].font = Font(bold=True, name="Arial")

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 40

    os.makedirs(OUTPUT_EXCEL.parent, exist_ok=True)
    wb.save(str(OUTPUT_EXCEL))
    log.info(f"Excel saved to {OUTPUT_EXCEL}")


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    log = setup_logging()
    start_time = time.time()

    log.info("=" * 60)
    log.info("DDDS Data Collector")
    log.info(f"Anchor: {ANCHOR_DATE}  |  Window: {WINDOW_START} to {WINDOW_END}")
    log.info("=" * 60)

    companies = build_company_universe(log)
    if not companies:
        log.error("No companies found. Exiting.")
        return
    log.info(f"Universe size: {len(companies)} companies")
    log.info(f"Estimated filing download time: {len(companies) * 18 * REQUEST_DELAY / 60:.0f}–"
             f"{len(companies) * 25 * REQUEST_DELAY / 60:.0f} minutes")

    summary = collect_all_filings(companies, log)

    generate_excel(summary, log)

    elapsed = time.time() - start_time
    complete = sum(1 for s in summary if s["completeness"] == "Complete")
    total_8ks = sum(s["8k_count"] for s in summary)
    log.info("=" * 60)
    log.info(f"Done! {len(summary)} companies processed.")
    log.info(f"  Complete filings: {complete}")
    log.info(f"  Total 8-Ks downloaded: {total_8ks}")
    log.info(f"  Total HTTP requests: {get_request_count()}")
    log.info(f"  Elapsed time: {elapsed/60:.1f} minutes")
    log.info(f"  Files: {BASE_DIR}/")
    log.info(f"  Excel: {OUTPUT_EXCEL}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()