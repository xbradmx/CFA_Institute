"""
DDDS Transcript Extractor
=========================
Processes 8-K filings already downloaded by ddds_data_collector.py to
find earnings call transcripts in EX-99 exhibits.

Pipeline:
  1. Load company universe from cache
  2. For each company, look at their 8-K folder
  3. For each 8-K, fetch the EDGAR filing index page
  4. Check if the 8-K covers Item 2.02 or 7.01 (earnings-related)
  5. If yes, find EX-99.x exhibit URLs from the index
  6. Download the exhibits
  7. Classify: transcript vs press release (keyword + structure check)
  8. Save transcripts to {TICKER}/Earnings_Transcripts/
  9. Update the Excel summary with a transcript status column

Usage:
  pip install requests openpyxl tqdm
  python ddds_transcript_extractor.py

Run AFTER ddds_data_collector.py has finished.
"""

import os
import re
import json
import time
import logging
import requests
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# ── Configuration ──────────────────────────────────────────────────────────

# !! Must match ddds_data_collector.py !!
USER_AGENT = "connorokeeffe07@gmail.com"

BASE_DIR = Path("DDDS/Companies")
EXCEL_PATH = Path("DDDS/company_filing_summary.xlsx")
CACHE_FILE = Path("DDDS/company_universe_cache.json")
LOG_FILE = Path("DDDS/ddds_transcript_extractor.log")

HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept-Encoding": "gzip, deflate",
}

REQUEST_DELAY = 0.15


# ── Transcript classification keywords ─────────────────────────────────────

# If enough of these appear in the document, it's a transcript
TRANSCRIPT_MARKERS = [
    "corporate participants",
    "conference call participants",
    "conference participants",
    "operator:",
    "prepared remarks",
    "question-and-answer",
    "question and answer",
    "q&a session",
    "open the line for questions",
    "open it up for questions",
    "we'll now take questions",
    "we will now take questions",
    "earnings conference call",
    "earnings call transcript",
    "good morning, everyone, and welcome to",
    "good afternoon, everyone, and welcome to",
    "good morning, and welcome to the",
    "good afternoon, and welcome to the",
    "before we begin, i'd like to remind",
    "before we begin, i would like to remind",
    "i would now like to turn the call over",
    "i'd now like to turn the call over",
    "opening remarks",
]

# Minimum marker hits to classify as transcript
MIN_MARKER_HITS = 3


# ── Setup ──────────────────────────────────────────────────────────────────

def setup_logging():
    os.makedirs("DDDS", exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(str(LOG_FILE)),
            logging.StreamHandler(),
        ],
    )
    return logging.getLogger("ddds_transcripts")


# ── HTTP ───────────────────────────────────────────────────────────────────

_request_count = 0


def rate_limit():
    time.sleep(REQUEST_DELAY)


def safe_get(url, max_retries=3):
    global _request_count
    for attempt in range(max_retries):
        try:
            rate_limit()
            _request_count += 1
            resp = requests.get(url, headers=HEADERS, timeout=30)
            if resp.status_code == 200:
                return resp
            elif resp.status_code == 429:
                wait = 15 * (attempt + 1)
                logging.warning(f"Rate limited. Waiting {wait}s...")
                time.sleep(wait)
                continue
            elif resp.status_code == 404:
                return None
            else:
                if attempt < max_retries - 1:
                    time.sleep(3)
                return None
        except requests.exceptions.RequestException as e:
            logging.warning(f"Request error: {e}")
            if attempt < max_retries - 1:
                time.sleep(5)
    return None


def strip_html(text):
    """Remove HTML tags and decode entities. Simple regex approach."""
    text = re.sub(r'<style[^>]*>.*?</style>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'&lt;', '<', text)
    text = re.sub(r'&gt;', '>', text)
    text = re.sub(r'&#\d+;', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


# ── Core Logic ─────────────────────────────────────────────────────────────

def get_filing_index(cik, accession_path):
    """
    Fetch the EDGAR filing index page for a given 8-K.
    Returns the HTML text, or None on failure.
    """
    cik_padded = str(cik).zfill(10)
    url = f"https://www.sec.gov/Archives/edgar/data/{cik_padded}/{accession_path}/"
    return safe_get(url)


def parse_index_page(index_html):
    """
    Parse a filing index page to extract:
    - Which items the 8-K covers (e.g. "2.02", "7.01")
    - URLs of EX-99.x exhibit files

    Returns: (set of item numbers, list of exhibit URLs)
    """
    text_lower = index_html.lower()

    # Extract items — look for "item 2.02", "item 7.01" etc.
    items = set()
    item_pattern = re.findall(r'item\s+(\d+\.\d+)', text_lower)
    for item in item_pattern:
        items.add(item)

    # Extract EX-99 exhibit URLs from the index table
    # The index page has rows like: EX-99.1  |  exhibit99-1.htm  |  ...
    exhibits = []
    # Match href links that look like exhibit files
    href_pattern = re.findall(
        r'href="([^"]*?)"[^>]*>',
        index_html,
        re.IGNORECASE
    )
    for href in href_pattern:
        href_lower = href.lower()
        # Match EX-99 type exhibits
        if ('ex-99' in href_lower or 'ex99' in href_lower or
                'exhibit99' in href_lower or 'exhibit-99' in href_lower):
            exhibits.append(href)

    # Also check the text for exhibit references with filenames
    # Some index pages list exhibits differently
    if not exhibits:
        row_pattern = re.findall(
            r'(?:ex-?99|exhibit\s*99)[^<]*?href="([^"]*?)"',
            index_html,
            re.IGNORECASE
        )
        exhibits.extend(row_pattern)

    return items, exhibits


def is_earnings_related(items):
    """Check if the 8-K items include earnings-related items."""
    earnings_items = {"2.02", "7.01"}
    return bool(items & earnings_items)


def classify_document(text):
    """
    Classify a document as transcript or press release.
    Returns: ("transcript", hit_count) or ("press_release", hit_count)
    """
    text_lower = text.lower()
    hits = 0
    for marker in TRANSCRIPT_MARKERS:
        if marker in text_lower:
            hits += 1

    # Also check for speaker turn patterns: "Name:" at start of lines
    # Transcripts have many of these, press releases don't
    speaker_turns = len(re.findall(r'\n\s*[A-Z][a-zA-Z\s\.\-]+:', text))

    if hits >= MIN_MARKER_HITS or (hits >= 2 and speaker_turns >= 10):
        return "transcript", hits
    else:
        return "press_release", hits


def process_company(cik, info, log):
    """
    Process all 8-Ks for a single company.
    Returns: dict with transcript extraction results.
    """
    ticker = info["ticker"].upper()
    eightk_dir = BASE_DIR / ticker / "8-K"
    transcript_dir = BASE_DIR / ticker / "Earnings_Transcripts"
    os.makedirs(transcript_dir, exist_ok=True)

    results = {
        "total_8ks": 0,
        "earnings_8ks": 0,
        "exhibits_found": 0,
        "transcripts_found": 0,
        "press_releases_found": 0,
        "transcript_dates": [],
    }

    if not eightk_dir.exists():
        return results

    # List all 8-K files
    eightk_files = sorted(eightk_dir.glob("8-K*"))
    results["total_8ks"] = len(eightk_files)

    for filepath in eightk_files:
        # Extract accession number from filename
        # Format: 8-K_2025-01-15_0001234567-25-000001.html
        fname = filepath.stem
        parts = fname.split("_")
        if len(parts) < 3:
            continue

        date_str = parts[1]
        accession_formatted = parts[2]
        accession_path = accession_formatted.replace("-", "")

        # Fetch the filing index page
        resp = get_filing_index(cik, accession_path)
        if not resp:
            continue

        index_html = resp.text

        # Parse for items and exhibits
        items, exhibit_urls = parse_index_page(index_html)

        # Only process earnings-related 8-Ks
        if not is_earnings_related(items):
            continue

        results["earnings_8ks"] += 1

        if not exhibit_urls:
            continue

        results["exhibits_found"] += 1

        # Download and classify each exhibit
        cik_padded = str(cik).zfill(10)
        for exhibit_url in exhibit_urls:
            # Build full URL if relative
            if exhibit_url.startswith("/"):
                full_url = f"https://www.sec.gov{exhibit_url}"
            elif not exhibit_url.startswith("http"):
                full_url = (
                    f"https://www.sec.gov/Archives/edgar/data/"
                    f"{cik_padded}/{accession_path}/{exhibit_url}"
                )
            else:
                full_url = exhibit_url

            exhibit_resp = safe_get(full_url)
            if not exhibit_resp or len(exhibit_resp.content) < 200:
                continue

            # Strip HTML to get plain text for classification
            raw_text = exhibit_resp.text
            plain_text = strip_html(raw_text)

            doc_type, hit_count = classify_document(plain_text)

            if doc_type == "transcript":
                results["transcripts_found"] += 1
                results["transcript_dates"].append(date_str)

                # Save the transcript (keep original HTML for structure)
                safe_date = date_str.replace("-", "")
                out_filename = f"transcript_{date_str}_{accession_formatted}.html"
                out_path = transcript_dir / out_filename
                if not out_path.exists():
                    with open(out_path, "wb") as f:
                        f.write(exhibit_resp.content)

                # Also save a plain text version
                txt_filename = f"transcript_{date_str}_{accession_formatted}.txt"
                txt_path = transcript_dir / txt_filename
                if not txt_path.exists():
                    with open(txt_path, "w", encoding="utf-8") as f:
                        f.write(plain_text)

                log.info(f"    TRANSCRIPT found: {date_str} "
                         f"({hit_count} markers matched)")
                break  # One transcript per 8-K is enough

            else:
                results["press_releases_found"] += 1

    return results


# ── Excel Update ───────────────────────────────────────────────────────────

def update_excel(transcript_data, log):
    """
    Update the existing Excel summary with transcript status columns.
    Adds columns after the existing data.
    """
    if not EXCEL_PATH.exists():
        log.error(f"Excel file not found: {EXCEL_PATH}")
        log.error("Run ddds_data_collector.py first.")
        return

    wb = load_workbook(str(EXCEL_PATH))
    ws = wb["Filing Summary"]

    # Find the current last column
    max_col = ws.max_column

    # Check if we've already added transcript columns (idempotent)
    existing_headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]
    if "EC Transcript" in existing_headers:
        # Find the column index and overwrite
        tc_col = existing_headers.index("EC Transcript") + 1
        td_col = tc_col + 1
        te_col = tc_col + 2
        log.info("Transcript columns already exist — updating in place.")
    else:
        # Add new columns
        tc_col = max_col + 1  # EC Transcript (Yes/No)
        td_col = max_col + 2  # Transcript Dates
        te_col = max_col + 3  # Transcript Count

    # Styles
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")

    # Write headers
    for col, (header, width) in enumerate(
        [("EC Transcript", 16), ("Transcript Dates", 30), ("Transcripts Found", 16)],
        start=tc_col
    ):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Write data for each row
    for row in range(2, ws.max_row + 1):
        ticker = ws.cell(row=row, column=1).value
        if not ticker:
            continue

        ticker = ticker.strip().upper()
        data = transcript_data.get(ticker, {})
        count = data.get("transcripts_found", 0)
        dates = "; ".join(sorted(data.get("transcript_dates", [])))
        has_transcript = count > 0

        # EC Transcript column
        cell = ws.cell(row=row, column=tc_col,
                       value="Yes" if has_transcript else "No")
        cell.border = border
        cell.font = Font(name="Arial", size=10)
        cell.fill = green if has_transcript else red

        # Transcript Dates column
        cell = ws.cell(row=row, column=td_col, value=dates)
        cell.border = border
        cell.font = Font(name="Arial", size=10)

        # Transcript Count column
        cell = ws.cell(row=row, column=te_col, value=count)
        cell.border = border
        cell.font = Font(name="Arial", size=10)

    # Update auto-filter to include new columns
    last_col = get_column_letter(te_col)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    wb.save(str(EXCEL_PATH))
    log.info(f"Excel updated: {EXCEL_PATH}")


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    log = setup_logging()
    start_time = time.time()

    log.info("=" * 60)
    log.info("DDDS Transcript Extractor")
    log.info("=" * 60)

    # Load company universe
    if not CACHE_FILE.exists():
        log.error(f"Cache file not found: {CACHE_FILE}")
        log.error("Run ddds_data_collector.py first.")
        return

    with open(CACHE_FILE) as f:
        companies = json.load(f)
    log.info(f"Loaded {len(companies)} companies from cache.")

    # Process each company
    transcript_data = {}  # ticker → results
    total = len(companies)
    total_transcripts = 0
    total_earnings_8ks = 0

    for idx, (cik, info) in enumerate(companies.items()):
        ticker = info["ticker"].upper()
        log.info(f"[{idx+1}/{total}] {ticker}")

        results = process_company(cik, info, log)
        transcript_data[ticker] = results

        total_transcripts += results["transcripts_found"]
        total_earnings_8ks += results["earnings_8ks"]

        if results["transcripts_found"] > 0:
            log.info(f"  → {results['transcripts_found']} transcript(s), "
                     f"{results['earnings_8ks']} earnings 8-Ks, "
                     f"{results['press_releases_found']} press releases")

    # Update Excel
    log.info("Updating Excel summary...")
    update_excel(transcript_data, log)

    elapsed = time.time() - start_time
    companies_with_transcripts = sum(
        1 for d in transcript_data.values() if d["transcripts_found"] > 0
    )

    log.info("=" * 60)
    log.info(f"Done! {total} companies processed.")
    log.info(f"  Earnings-related 8-Ks found: {total_earnings_8ks}")
    log.info(f"  Transcripts extracted: {total_transcripts}")
    log.info(f"  Companies with transcripts: {companies_with_transcripts}/{total}")
    log.info(f"  HTTP requests: {_request_count}")
    log.info(f"  Elapsed: {elapsed/60:.1f} minutes")
    log.info("=" * 60)


if __name__ == "__main__":
    main()